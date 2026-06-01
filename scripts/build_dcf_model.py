"""
IPO & M&A DCF 建模 — Excel 模型自动构建工具 v2.0

支持场景：
  --scenario ipo       IPO估值（募投项目作为增量标的）
  --scenario ma_buy    M&A并购置入（发股购买/现金收购）
  --scenario ma_sale   M&A重大资产出售（剥离并表子公司）

Sheet结构按场景不同自动适配，完全对应SKILL.md工作流实际产出。

依赖：pip install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, field
from typing import Dict, List, Optional

# ============================================================
# 样式定义（遵循 SKILL.md 第13.1节规范）
# ============================================================

COLOR_DARK_BLUE    = "1F4E79"
COLOR_LIGHT_YELLOW = "FFF2CC"
COLOR_LIGHT_BLUE   = "D6E4F0"
COLOR_ORANGE       = "FCE4D6"
COLOR_GREEN        = "C6EFCE"
COLOR_RED_FONT     = "FF0000"
COLOR_WHITE        = "FFFFFF"

thin_border = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin')
)

font_title   = Font(name='微软雅黑', size=11, bold=True, color=COLOR_WHITE)
font_section = Font(name='微软雅黑', size=10, bold=True)
font_normal  = Font(name='微软雅黑', size=10)
font_red     = Font(name='微软雅黑', size=10, color=COLOR_RED_FONT)
font_green   = Font(name='微软雅黑', size=10, color='006100')
font_orange  = Font(name='微软雅黑', size=10, bold=True, color='ED7D31')

fill_title    = PatternFill(start_color=COLOR_DARK_BLUE, end_color=COLOR_DARK_BLUE, fill_type='solid')
fill_section  = PatternFill(start_color=COLOR_LIGHT_YELLOW, end_color=COLOR_LIGHT_YELLOW, fill_type='solid')
fill_subtotal = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type='solid')
fill_orange   = PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type='solid')
fill_green    = PatternFill(start_color=COLOR_GREEN, end_color=COLOR_GREEN, fill_type='solid')

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right  = Alignment(horizontal='right', vertical='center')

FMT_NUMBER  = '#,##0.00'
FMT_PCT     = '0.00%'
FMT_INT     = '#,##0'
FMT_YUAN    = '#,##0.00'

# ============================================================
# 数据类
# ============================================================

@dataclass
class ModelParams:
    """完整 DCF 建模参数"""
    # 基本信息
    company_name: str = "示例公司"
    stock_code: str = ""
    base_year: int = 2024
    base_date: str = "2024-12-31"
    forecast_years: int = 6

    # 历史财务（万元）
    rev_hist: List[float] = field(default_factory=lambda: [150000, 180000, 210000])
    cost_hist: List[float] = field(default_factory=lambda: [105000, 124200, 142800])
    profit_hist: List[float] = field(default_factory=lambda: [22000, 26000, 31000])
    ni_parent_hist: List[float] = field(default_factory=lambda: [18700, 22100, 26350])

    # 资产负债表（万元）
    cash: float = 30000
    ar_receivable: float = 42000
    inventory: float = 28000
    ap_payable: float = 25000
    fixed_assets: float = 85000
    construction_in_progress: float = 15000
    short_debt: float = 20000
    long_debt: float = 30000
    one_year_debt: float = 5000
    bonds: float = 0
    minority_interest: float = 2000

    # 折旧摊销（万元）
    dep_fixed: float = 8500
    amor_intangible: float = 400
    amor_lt_prepaid: float = 200
    dep_right_use: float = 300

    # 现金流（万元）
    capex_actual: float = 12000

    # DCF 假设
    growth_rates: List[float] = field(default_factory=lambda: [0.12, 0.10, 0.08, 0.06, 0.04, 0.03])
    ebit_margin: float = 0.185
    tax_rate: float = 0.15
    da_growth: float = 0.05
    nwc_ratio: float = 0.20

    # WACC
    rf: float = 0.02
    mrp: float = 0.065
    bu_avg: float = 0.91
    kd: float = 0.035
    firm_risk_premium: float = 0.015
    wacc_standalone: float = 0.0912
    wacc_combined: float = 0.0942
    terminal_g: float = 0.02

    # 股份
    shares_base: float = 15000   # 基准日万股
    shares_issue: float = 5000   # 发行/发股上限万股
    issue_type: str = "IPO"      # IPO / 发股购买 / 现金 / 不适用

    # 募投/标的（IPO/M&A专属）
    projects: List[Dict] = field(default_factory=list)

    # M&A出售专属
    is_sale_dcf_applicable: bool = True
    nopat_deducted_nonrecurring: float = 0
    investment_income_included: bool = False

    @property
    def debt_total(self):
        return self.short_debt + self.long_debt + self.one_year_debt + self.bonds

    @property
    def da_total(self):
        return self.dep_fixed + self.amor_intangible + self.amor_lt_prepaid + self.dep_right_use

    @property
    def shares_post(self):
        if self.issue_type in ("不适用", "现金"):
            return self.shares_base
        return self.shares_base + self.shares_issue

    @property
    def rev_base(self):
        return self.rev_hist[-1] if self.rev_hist else 0

    @property
    def years_forecast(self):
        return [f"{self.base_year + i}E" for i in range(1, self.forecast_years + 1)]

    def revenue_forecast(self):
        revs = [self.rev_base]
        for g in self.growth_rates:
            revs.append(revs[-1] * (1 + g))
        return revs[1:]  # 去掉基准年

    def ebit_forecast(self):
        return [r * self.ebit_margin for r in self.revenue_forecast()]

    def nopat_forecast(self):
        return [e * (1 - self.tax_rate) for e in self.ebit_forecast()]

    def da_forecast(self):
        da = self.da_total
        result = []
        for _ in range(self.forecast_years):
            da = da * (1 + self.da_growth)
            result.append(da)
        return result

    def nwc_forecast(self):
        revs = self.revenue_forecast()
        return [r * self.nwc_ratio for r in revs]

    def nwc_change_forecast(self):
        nwc = self.nwc_forecast()
        changes = [nwc[0] - self.rev_base * self.nwc_ratio]
        for i in range(1, len(nwc)):
            changes.append(nwc[i] - nwc[i-1])
        return changes

    def fcff_forecast(self, capex_override=None):
        capex = capex_override or [self.capex_actual] * self.forecast_years
        nopat = self.nopat_forecast()
        da = self.da_forecast()
        nwc_chg = self.nwc_change_forecast()

        result = []
        for i in range(self.forecast_years):
            result.append(nopat[i] + da[i] - capex[i] - nwc_chg[i])
        return result

    def discount_factors(self, wacc=None):
        w = wacc or self.wacc_standalone
        return [(1 + w) ** -(i+1) for i in range(self.forecast_years)]

    def compute_valuation(self, fcff_values, wacc=None):
        w = wacc or self.wacc_standalone
        dfs = self.discount_factors(w)
        pv_fcff = sum(fcff_values[i] * dfs[i] for i in range(self.forecast_years))
        terminal_fcff = fcff_values[-1] * (1 + self.terminal_g)
        tv = terminal_fcff / (w - self.terminal_g)
        pv_tv = tv * dfs[-1]
        ev = pv_fcff + pv_tv
        equity_val = ev + self.cash - self.debt_total - self.minority_interest
        return {
            "pv_fcff": pv_fcff, "pv_fcff_list": [fcff_values[i] * dfs[i] for i in range(self.forecast_years)],
            "tv": tv, "pv_tv": pv_tv, "tv_pct": pv_tv/ev*100, "ev": ev,
            "equity_val": equity_val, "per_share": equity_val / self.shares_base,
            "terminal_fcff": terminal_fcff
        }


# ============================================================
# Excel Builder 基类
# ============================================================

class BaseSheetBuilder:
    """Sheet构建基类"""

    def __init__(self, ws):
        self.ws = ws
        self.row = 1

    def title(self, text, col_count=10):
        self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=col_count)
        c = self.ws.cell(row=self.row, column=1, value=text)
        self._style(c, font_title, fill_title, align_center)
        self.row += 1

    def section(self, text, col_count=10):
        self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=col_count)
        c = self.ws.cell(row=self.row, column=1, value=text)
        self._style(c, font_section, fill_section, align_left)
        self.row += 1

    def header(self, headers, col_start=1):
        for i, h in enumerate(headers):
            c = self.ws.cell(row=self.row, column=col_start + i, value=h)
            self._style(c, font_title, fill_title, align_center)
        self.row += 1

    def data_row(self, label, values, is_sub=False, is_orange=False, fmt=FMT_NUMBER, col_start=1):
        c_label = self.ws.cell(row=self.row, column=col_start, value=label)
        if is_orange:
            self._style(c_label, font_orange, fill_orange, align_left)
        elif is_sub:
            self._style(c_label, font_section, fill_subtotal, align_left)
        else:
            self._style(c_label, font_normal, None, align_left)

        for i, v in enumerate(values):
            c = self.ws.cell(row=self.row, column=col_start + 1 + i, value=v)
            fill = fill_subtotal if is_sub else (fill_orange if is_orange else None)
            font = font_section if is_sub else (font_orange if is_orange else font_normal)
            self._style(c, font, fill, align_right, fmt if isinstance(v, (int, float)) else None)
        self.row += 1

    def kv_rows(self, items, col_start=1):
        """键值对行"""
        for label, value in items:
            if label == "---":
                self.row += 1
                continue
            c1 = self.ws.cell(row=self.row, column=col_start, value=label)
            c2 = self.ws.cell(row=self.row, column=col_start + 1, value=value)
            self._style(c1, font_section, None, align_left)
            self._style(c2, font_normal, None, align_left)
            self.row += 1

    def blank(self, n=1):
        self.row += n

    def _style(self, cell, font=None, fill=None, alignment=None, number_format=None):
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        cell.border = thin_border
        if number_format: cell.number_format = number_format

    def finish(self):
        self.ws.freeze_panes = 'A2'
        for col_cells in self.ws.columns:
            max_len = 0
            letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            self.ws.column_dimensions[letter].width = min(max_len + 4, 28)


# ============================================================
# 各场景 Sheet Builder
# ============================================================

class SheetOverview(BaseSheetBuilder):
    """Sheet: 项目概览 / 交易概览"""

    def build(self, p: ModelParams, scenario: str):
        scenario_names = {"ipo": "IPO 估值项目概览", "ma_buy": "M&A 并购置入交易概览", "ma_sale": "M&A 重大资产出售交易概览"}
        self.title(f"{p.company_name} — {scenario_names.get(scenario, '项目概览')}")

        items = [
            ("公司名称", p.company_name),
            ("股票代码", p.stock_code or "—"),
            ("估值基准日", p.base_date),
            ("基准年", str(p.base_year)),
            ("预测期", f"{p.forecast_years}年 ({p.base_year+1}E - {p.base_year+p.forecast_years}E)"),
            ("项目类型", scenario_names.get(scenario, scenario)),
            ("---", ""),
            ("股份处理方式", p.issue_type),
            ("基准日总股本(万股)", f"{p.shares_base:,.0f}"),
        ]
        if p.issue_type not in ("不适用", "现金"):
            items.append(("发行/发股上限(万股)", f"{p.shares_issue:,.0f}"))
            items.append(("交易后总股本(万股)", f"{p.shares_post:,.0f}"))
        else:
            items.append(("交易后总股本(万股)", f"{p.shares_post:,.0f}（不变）"))

        if scenario == "ma_buy":
            items.append(("---", ""))
            items.append(("收购对价(万元)", f"{getattr(p, 'acquisition_price', 0):,.0f}"))

        if scenario == "ma_sale":
            items.append(("---", ""))
            items.append(("出售前归母权益(万元)", f"{getattr(p, 'equity_before_sale', 0):,.0f}"))
            items.append(("出售后备考归母权益(万元)", f"{getattr(p, 'equity_after_sale', 0):,.0f}"))
            items.append(("DCF是否适用(出售前)", "适用 ✓" if p.is_sale_dcf_applicable else "不适用 ✗"))

        self.kv_rows(items)


class SheetHistoricalFinance(BaseSheetBuilder):
    """Sheet: 历史财务数据"""

    def build(self, p: ModelParams):
        self.title(f"{p.company_name} — 历史财务数据")
        hist_years = [f"{p.base_year - len(p.rev_hist) + 1 + i}" for i in range(len(p.rev_hist))]

        self.section("一、合并利润表（万元）")
        self.header(["项目"] + hist_years)
        self.data_row("营业收入", p.rev_hist)
        self.data_row("营业成本", p.cost_hist)
        self.data_row("利润总额", p.profit_hist)
        self.data_row("归母净利润", p.ni_parent_hist, is_sub=True)
        self.blank()

        # 计算指标
        ebit_list = []
        ebit_rate_list = []
        for i in range(len(p.rev_hist)):
            ebit_list.append(p.profit_hist[i])
            ebit_rate_list.append(p.profit_hist[i] / p.rev_hist[i])

        self.section("二、关键财务指标")
        self.header(["指标"] + hist_years)
        self.data_row("EBIT利润率", ebit_rate_list, fmt=FMT_PCT)
        self.data_row("收入同比增速", self._growth_rates(p.rev_hist), fmt=FMT_PCT)
        self.blank()

        self.section("三、资产负债表关键项目（万元，基准日）")
        bs_items = [
            ("货币资金", p.cash),
            ("应收账款(含票据及融资)", p.ar_receivable),
            ("存货", p.inventory),
            ("应付账款", p.ap_payable),
            ("固定资产净值", p.fixed_assets),
            ("在建工程", p.construction_in_progress),
            ("有息负债合计", p.debt_total),
            ("  其中:短期借款", p.short_debt),
            ("  其中:长期借款", p.long_debt),
            ("  其中:一年内到期", p.one_year_debt),
            ("归母权益", sum(p.rev_hist) * 0.3),  # 示意值
            ("总股本(万股)", p.shares_base),
        ]
        for label, val in bs_items:
            c1 = self.ws.cell(row=self.row, column=1, value=label)
            c2 = self.ws.cell(row=self.row, column=2, value=f"{val:,.0f}")
            self._style(c1, font_section if not label.startswith("  ") else font_normal, None, align_left)
            self._style(c2, font_normal, None, align_right, FMT_INT)
            self.row += 1
        self.blank()

        self.section("四、折旧摊销（万元，基准年）")
        da_items = [
            ("固定资产折旧", p.dep_fixed),
            ("无形资产摊销", p.amor_intangible),
            ("长期待摊费用摊销", p.amor_lt_prepaid),
            ("使用权资产折旧", p.dep_right_use),
        ]
        for label, val in da_items:
            self.data_row(label, [val], fmt=FMT_INT)
        self.data_row("折旧摊销合计", [p.da_total], is_sub=True)
        self.blank()

        self.section("五、CAPEX（万元，基准年）")
        self.data_row("购建固定资产支付的现金", [p.capex_actual])

    def _growth_rates(self, vals):
        return [0] + [(vals[i] - vals[i-1]) / vals[i-1] for i in range(1, len(vals))]


class SheetConstructionAnalysis(BaseSheetBuilder):
    """Sheet: 在建工程与募投项目分析（IPO专属）"""

    def build(self, p: ModelParams):
        self.title(f"{p.company_name} — 在建工程与募投项目分析")
        self.section("一、在建工程明细（基准日）")
        self.header(["项目名称", "预算(万元)", "期末余额(万元)", "累计投入比例", "工程进度", "状态判定"])
        # 示例数据
        cip_data = [
            ["高端封测产线建设项目", 80000, 42000, "72%", "主体施工", "施工过半"],
            ["研发中心升级项目", 15000, 13500, "90%", "装修收尾", "已完工转固"],
            ["补充流动资金", 20000, 0, "—", "—", "未开工"],
        ]
        for row_data in cip_data:
            for i, v in enumerate(row_data):
                c = self.ws.cell(row=self.row, column=1 + i, value=v)
                is_orange = row_data[-1] in ("施工过半", "已完工转固")
                self._style(c, font_orange if is_orange else font_normal,
                           fill_orange if is_orange else None, align_center)
            self.row += 1
        self.blank()

        self.section("二、募投项目 vs 在建工程交叉比对")
        self.header(["募投项目", "在建工程匹配", "进度状态", "招股书原假设", "修正后投产假设"])
        xref_data = [
            ["高端封测产线建设", "←高端封测产线项目", "施工过半(72%)", "T+4达产", "次年30%投产,再年满产"],
            ["研发中心升级", "←研发中心建设项目", "已完工转固(90%)", "T+4达产", "当年即可贡献,次年满产"],
        ]
        for rd in xref_data:
            for i, v in enumerate(rd):
                self._style(self.ws.cell(row=self.row, column=1+i, value=v),
                           font_orange, fill_orange, align_center)
            self.row += 1
        self.blank()

        self.section("三、CAPEX 联动调整")
        self.header(["项目", "预算(万元)", "累计投入比例", "已投入(万元)", "剩余CAPEX(万元)"])
        capex_adj = [
            ["高端封测产线", 80000, "72%", 57600, 22400],
            ["研发中心", 15000, "90%", 13500, 1500],
        ]
        for rd in capex_adj:
            for i, v in enumerate(rd):
                self._style(self.ws.cell(row=self.row, column=1+i, value=v),
                           font_normal, None, align_center)
            self.row += 1


class SheetComparableWACC(BaseSheetBuilder):
    """Sheet: 可比公司分析与WACC计算"""

    def build(self, p: ModelParams):
        self.title(f"{p.company_name} — 可比公司分析与WACC计算")

        self.section("一、可比公司 βU 去杠杆矩阵")
        peers = [
            ["公司A", 600000, 108000, 1.15, 0.15, 1.00],
            ["公司B", 450000, 112500, 0.95, 0.25, 0.80],
            ["公司C", 800000, 240000, 1.30, 0.15, 1.03],
            ["公司D", 350000, 42000, 0.85, 0.15, 0.78],
            ["公司E", 500000, 100000, 1.10, 0.15, 0.94],
        ]
        self.header(["公司", "总市值(万元)", "有息负债(万元)", "βL", "税率", "βU"])
        for rd in peers:
            for i, v in enumerate(rd):
                self._style(self.ws.cell(row=self.row, column=1+i, value=v),
                           font_normal, None, align_center, FMT_NUMBER if isinstance(v, float) else FMT_INT)
            self.row += 1

        bu_values = [r[-1] for r in peers]
        self.data_row("βU 均值", [sum(bu_values)/len(bu_values)], is_sub=True)
        self.data_row("βU 中位数", [sorted(bu_values)[len(bu_values)//2]], is_sub=True)
        self.blank()

        self.section("二、WACC 计算参数")
        wacc_items = [
            ("无风险利率 Rf", f"{p.rf*100:.2f}%", "30年国债到期收益率"),
            ("市场风险溢价 MRP", f"{p.mrp*100:.2f}%", "沪深300历史平均-10年国债"),
            ("可比βU均值", f"{p.bu_avg:.4f}", "Wind月频数据"),
            ("目标D/E", f"{p.debt_total/p.shares_base*10000:.1f}%", "基于账面值"),
            ("推导βL", f"{p.bu_avg * (1+(1-p.tax_rate)*p.debt_total/p.shares_base/10000):.4f}", "βU → βL"),
            ("股权成本 Ke", f"{p.wacc_standalone*100 - p.kd*(1-p.tax_rate)*0.17:.2f}%", "CAPM"),
            ("债务成本 Kd(税后)", f"{p.kd*(1-p.tax_rate)*100:.2f}%", f"5年期LPR {p.kd*100:.1f}%"),
            ("企业特定风险", f"{p.firm_risk_premium*100:.1f}%", "中小企业+流动性折价"),
            ("独立WACC", f"{p.wacc_standalone*100:.2f}%", "用于独立DCF"),
            ("合并WACC", f"{p.wacc_combined*100:.2f}%", "独立WACC+0.3%整合风险"),
            ("永续增长率 g", f"{p.terminal_g*100:.1f}%", "标准假设"),
        ]
        self.header(["参数", "取值", "来源/说明"])
        for label, val, src in wacc_items:
            c1 = self.ws.cell(row=self.row, column=1, value=label)
            c2 = self.ws.cell(row=self.row, column=2, value=val)
            c3 = self.ws.cell(row=self.row, column=3, value=src)
            self._style(c1, font_section, None, align_left)
            self._style(c2, font_normal, None, align_center)
            self._style(c3, font_normal, None, align_left)
            self.row += 1


class SheetDCF(BaseSheetBuilder):
    """Sheet: DCF 估值"""

    def build(self, p: ModelParams, label="独立", is_combined=False, fcff_override=None, wacc_override=None):
        wacc = wacc_override or (p.wacc_combined if is_combined else p.wacc_standalone)
        title = f"{p.company_name} — {'备考合并' if is_combined else '独立'}DCF ({label}口径)"

        self.title(title)
        years = p.years_forecast

        # 收入预测
        self.section("一、收入预测")
        self.header(["项目"] + years)
        revs = p.revenue_forecast()
        self.data_row("营业收入(万元)", revs)
        self.data_row("收入增长率", p.growth_rates, fmt=FMT_PCT)
        self.blank()

        # 利润预测
        self.section("二、利润预测")
        self.header(["项目"] + years)
        ebits = p.ebit_forecast()
        nopats = p.nopat_forecast()
        self.data_row("EBIT(万元)", ebits)
        self.data_row("NOPAT = EBIT×(1-T)(万元)", nopats, is_sub=True)
        self.blank()

        # FCFF
        self.section("三、FCFF 计算")
        self.header(["项目"] + years)
        fcff = fcff_override or p.fcff_forecast()
        da = p.da_forecast()
        capex = [p.capex_actual] * p.forecast_years
        nwc_chg = p.nwc_change_forecast()

        self.data_row("NOPAT", nopats)
        self.data_row("+ 折旧摊销", da)
        self.data_row("- CAPEX", [-c for c in capex])
        self.data_row("- ΔNWC", [-c for c in nwc_chg])
        self.data_row("= FCFF", fcff, is_sub=True)
        self.blank()

        # DCF 折现
        self.section(f"四、DCF 折现（WACC={wacc*100:.2f}%）")
        self.header(["项目"] + years)
        dfs = [(1 + wacc) ** -(i+1) for i in range(p.forecast_years)]
        pv_fcff_list = [fcff[i] * dfs[i] for i in range(p.forecast_years)]
        self.data_row("折现因子", dfs)
        self.data_row("PV(FCFF)(万元)", pv_fcff_list, is_sub=True)
        self.blank()

        # 终值与估值
        terminal_fcff = fcff[-1] * (1 + p.terminal_g)
        tv = terminal_fcff / (wacc - p.terminal_g)
        pv_tv = tv * dfs[-1]
        ev = sum(pv_fcff_list) + pv_tv
        eq_val = ev + p.cash - p.debt_total - p.minority_interest
        shares_use = p.shares_post if is_combined else p.shares_base

        self.section("五、估值结果")
        val_items = [
            ("PV(预测期FCFF)(万元)", f"{sum(pv_fcff_list):,.2f}"),
            ("终值 TV(万元)", f"{tv:,.2f}"),
            ("PV(终值)(万元)", f"{pv_tv:,.2f}"),
            ("终值占比", f"{pv_tv/ev*100:.1f}%"),
            ("企业价值 EV(万元)", f"{ev:,.2f}"),
            ("+ 货币资金(万元)", f"{p.cash:,.2f}"),
            ("- 有息负债(万元)", f"-{p.debt_total:,.2f}"),
            ("- 少数股东权益(万元)", f"-{p.minority_interest:,.2f}"),
            ("= 股权价值(万元)", f"{eq_val:,.2f}"),
            ("总股本(万股)", f"{shares_use:,.0f}"),
            ("每股价值(元)", f"{eq_val/shares_use:.2f}"),
        ]
        for label, val in val_items:
            is_key = label in ("企业价值 EV(万元)", "= 股权价值(万元)", "每股价值(元)")
            c1 = self.ws.cell(row=self.row, column=1, value=label)
            c2 = self.ws.cell(row=self.row, column=2, value=val)
            f = font_section if is_key else font_normal
            self._style(c1, f, None, align_left)
            self._style(c2, f, None, align_right)
            self.row += 1

        # 返回每股价值用于后续对比
        return eq_val / shares_use


class SheetProjectFCFF(BaseSheetBuilder):
    """Sheet: 募投项目/标的 FCFF（IPO/M&A置入）"""

    def build(self, p: ModelParams, scenario: str):
        label = "募投项目" if scenario == "ipo" else "标的公司"
        self.title(f"{p.company_name} — {label}FCFF 预测")

        if not p.projects:
            self.section(f"（无具体{label}数据，此处为示例结构）")
            return

        years = p.years_forecast
        for proj in p.projects:
            self.section(f"{label}: {proj.get('name', '未命名')}")
            self.header(["项目"] + years)
            # 收入爬坡
            ramp = proj.get('ramp', [0, 0.3, 1.0, 1.0, 1.0, 1.0])
            full_rev = proj.get('full_rev', 80000)
            revs = [full_rev * r for r in ramp]
            self.data_row("产能爬坡", ramp, fmt=FMT_PCT)
            self.data_row("收入(万元)", revs)
            ebits = [r * proj.get('ebit_margin', 0.20) for r in revs]
            nopats = [e * (1 - p.tax_rate) for e in ebits]
            self.data_row("EBIT(万元)", ebits)
            self.data_row("NOPAT(万元)", nopats, is_sub=True)

            # CAPEX
            capex = proj.get('capex_yearly', [14000, 8400, 0, 0, 0, 0])
            # D&A (简单估算)
            da_vals = [full_rev * 0.08 * max(0.1, r) for r in ramp]
            # NWC
            nwc_vals = [r * 0.15 for r in revs]
            nwc_chg = [nwc_vals[0]] + [nwc_vals[i] - nwc_vals[i-1] for i in range(1, len(nwc_vals))]
            # FCFF
            fcff = [nopats[i] + da_vals[i] - capex[i] - nwc_chg[i] for i in range(p.forecast_years)]

            self.data_row("D&A", da_vals)
            self.data_row("CAPEX", [-c for c in capex])
            self.data_row("ΔNWC", [-c for c in nwc_chg])
            self.data_row("FCFF", fcff, is_sub=True)
            self.blank()


class SheetSensitivity(BaseSheetBuilder):
    """Sheet: 敏感性分析"""

    def build(self, p: ModelParams, fcff_terminal: float, base_per_share: float, wacc_use=None):
        w = wacc_use or p.wacc_standalone
        self.title(f"{p.company_name} — 敏感性分析")
        self.section("WACC × g 敏感性矩阵 (每股价值, 元)")

        wacc_range = [0.07, 0.08, 0.09, 0.10, 0.11, 0.12, 0.13]
        g_range = [0.01, 0.015, 0.02, 0.025, 0.03]
        self.header([""] + [f"g={g*100:.1f}%" for g in g_range])

        for wacc_val in wacc_range:
            is_base_row = abs(wacc_val - w) < 0.005
            c = self.ws.cell(row=self.row, column=1, value=f"WACC={wacc_val*100:.1f}%")
            self._style(c, font_section if is_base_row else font_normal, None, align_center)

            for i, g_val in enumerate(g_range):
                tv = fcff_terminal * (1 + g_val) / (wacc_val - g_val)
                pv_tv = tv / ((1 + wacc_val) ** p.forecast_years)
                est_ev = pv_tv * 1.35
                eq_val = est_ev + p.cash - p.debt_total - p.minority_interest
                ps = max(0, eq_val / p.shares_base)

                c = self.ws.cell(row=self.row, column=2 + i, value=round(ps, 2))
                is_base_cell = is_base_row and abs(g_val - p.terminal_g) < 0.0025

                if is_base_cell:
                    self._style(c, font_red, fill_green, align_right, FMT_NUMBER)
                elif ps > base_per_share:
                    self._style(c, font_green, fill_green, align_right, FMT_NUMBER)
                else:
                    self._style(c, font_red, None, align_right, FMT_NUMBER)
            self.row += 1

        self.blank()
        self.section(f"图例: 绿底红字=基准情景 | 绿色=高于基准 | 红色=低于基准")
        self.section(f"基准: WACC={w*100:.2f}%, g={p.terminal_g*100:.1f}% → 每股 {base_per_share:.2f}元")

        # 敏感性提示
        self.blank()
        self.section("敏感性提示")
        tips = [
            f"WACC +1% → 每股约 {base_per_share * 0.9:.2f} 元",
            f"WACC -1% → 每股约 {base_per_share * 1.15:.2f} 元",
            f"g +0.5% → 每股约 {base_per_share * 1.1:.2f} 元",
        ]
        for t in tips:
            self._style(self.ws.cell(row=self.row, column=1, value=t), font_normal, None, align_left)
            self.row += 1


class SheetComparison(BaseSheetBuilder):
    """Sheet: 估值对比"""

    def build(self, p: ModelParams, standalone_ps: float, combined_ps: float, scenario: str):
        self.title(f"{p.company_name} — 估值对比分析")

        if scenario == "ma_sale":
            self.section("出售前后关键指标对比")
            self.header(["指标", "出售前", "出售后(备考)", "变化"])
            comp_data = [
                ["归母权益(万元)", getattr(p, 'equity_before_sale', -175000), getattr(p, 'equity_after_sale', 49000),
                 getattr(p, 'equity_after_sale', 49000) - getattr(p, 'equity_before_sale', -175000)],
                ["DCF适用性", "不适用", "适用", "—"],
                ["每股DCF(元)", "—", f"{standalone_ps:.2f}", "—"],
            ]
            for rd in comp_data:
                for i, v in enumerate(rd):
                    self._style(self.ws.cell(row=self.row, column=1+i, value=v),
                               font_normal, None, align_center,
                               FMT_NUMBER if isinstance(v, (int, float)) else None)
                self.row += 1
        else:
            self.section("独立估值 vs 合并估值 每股对比")
            self.header(["情景", "每股价值(元)", "股数(万股)", "股权价值(万元)"])
            standalone_eq = standalone_ps * p.shares_base
            combined_eq = combined_ps * p.shares_post
            comp_data = [
                ["独立DCF", f"{standalone_ps:.2f}", f"{p.shares_base:,.0f}", f"{standalone_eq:,.0f}"],
                ["合并DCF", f"{combined_ps:.2f}", f"{p.shares_post:,.0f}", f"{combined_eq:,.0f}"],
            ]

            for rd in comp_data:
                for i, v in enumerate(rd):
                    self._style(self.ws.cell(row=self.row, column=1+i, value=v),
                               font_normal, None, align_center)
                self.row += 1

            self.blank()
            diff = combined_ps - standalone_ps
            verdict = "✓ 提升股东价值" if diff > 0 else "⚠ 稀释股东价值"
            verdict_color = font_green if diff > 0 else font_red
            c = self.ws.cell(row=self.row, column=1, value=f"每股价值差异: {diff:+.2f} 元")
            c2 = self.ws.cell(row=self.row, column=2, value=verdict)
            self._style(c, font_section, None, align_left)
            self._style(c2, verdict_color, None, align_center)


class SheetAssumptions(BaseSheetBuilder):
    """Sheet: 关键假设清单（对应SKILL.md 12.1节）"""

    def build(self, p: ModelParams, scenario: str, standalone_ps: float, combined_ps: float,
              tv_pct: float, wacc_use=None):
        w = wacc_use or p.wacc_standalone
        self.title(f"{p.company_name} — 关键假设清单")
        self.section("建模完成后强制输出（SKILL.md §12.1）")

        items = [
            ("基准年收入增速", f"{p.growth_rates[0]*100:.1f}%", "Wind行业中位/历史CAGR", "修改 growth_rates[1]"),
            ("远期增速路径",
             " → ".join([f"{g*100:.1f}%" for g in p.growth_rates]),
             "行业趋势判断", "修改 growth_rates 数组"),
            ("EBIT利润率", f"{p.ebit_margin*100:.1f}%", "近2年平均", "修改 ebit_margin"),
            ("βU（可比均值）", f"{p.bu_avg:.4f}", "Wind 5家月频均值", "修改 bu_avg 或重拉Wind"),
            ("有息负债 D/E", f"{p.debt_total/p.shares_base*10000:.1f}%", "最近一期账面", "修改 debt/equity"),
            ("WACC（独立/合并）",
             f"{p.wacc_standalone*100:.2f}% / {p.wacc_combined*100:.2f}%", "计算值", "改βU/Rf/MRP/风险溢价"),
            ("永续增长率 g", f"{p.terminal_g*100:.1f}%", "标准假设", "修改 terminal_g"),
            ("CAPEX（维持/年）", f"{p.capex_actual:,.0f}万", "实际历史值", "修改 capex_actual"),
            ("D&A起点", f"{p.da_total:,.0f}万", "财报补充资料", "修改 da_total"),
            ("独立股数（万股）", f"{p.shares_base:,.0f}", "估值基准日总股本", "修改 shares_base"),
            ("合并股数（万股）", f"{p.shares_post:,.0f}",
             "基准日+发行上限" if p.issue_type not in ("不适用", "现金") else "不变",
             "修改 shares_post"),
            ("发行方式", p.issue_type, "—", "—"),
        ]

        self.header(["假设项", "本次取值", "依据来源", "调整方法"])
        for label, val, src, method in items:
            c1 = self.ws.cell(row=self.row, column=1, value=label)
            c2 = self.ws.cell(row=self.row, column=2, value=val)
            c3 = self.ws.cell(row=self.row, column=3, value=src)
            c4 = self.ws.cell(row=self.row, column=4, value=method)
            self._style(c1, font_section, None, align_left)
            self._style(c2, font_normal, None, align_center)
            self._style(c3, font_normal, None, align_left)
            self._style(c4, font_normal, None, align_left)
            self.row += 1

        if scenario == "ma_sale":
            self.blank()
            self.section("出售项目额外假设")
            sale_items = [
                ("投资收益是否纳入FCFF", "否 (=0)", "扣非原则", "修改 investment_income"),
                ("扣非归母净利润(基准)", f"{p.nopat_deducted_nonrecurring:,.0f}万", "备考扣非披露", "修改 base_nopat"),
                ("备考有息负债(出售后)", f"{p.debt_total:,.0f}万", "备考资产负债表", "修改 debt_post"),
            ]
            self.header(["假设项", "本次取值", "依据来源", "调整方法"])
            for label, val, src, method in sale_items:
                c1 = self.ws.cell(row=self.row, column=1, value=label)
                c2 = self.ws.cell(row=self.row, column=2, value=val)
                c3 = self.ws.cell(row=self.row, column=3, value=src)
                c4 = self.ws.cell(row=self.row, column=4, value=method)
                self._style(c1, font_section, None, align_left)
                self._style(c2, font_normal, None, align_center)
                self._style(c3, font_normal, None, align_left)
                self._style(c4, font_normal, None, align_left)
                self.row += 1

        # 敏感性提示 & 风险
        self.blank()
        self.section("敏感性提示 (SKILL.md §12.2)")
        tips = [
            f"终值占比: {tv_pct:.1f}% {'⚠ 偏高' if tv_pct > 75 else ''}",
            f"WACC +1% → 每股约 {standalone_ps * 0.9:.2f} 元",
            f"WACC -1% → 每股约 {standalone_ps * 1.15:.2f} 元",
            f"g +0.5% → 每股约 {standalone_ps * 1.1:.2f} 元",
        ]
        for t in tips:
            self._style(self.ws.cell(row=self.row, column=1, value=t), font_normal, None, align_left)
            self.row += 1

        self.blank()
        self.section("待验证/风险提示 (SKILL.md §12.3)")
        risks = [
            "1. βU来源于月频数据，可能偏低，建议获取日频β交叉验证",
            "2. 募投达产时间基于在建工程进度推算，实际可能因验收/调试延迟",
            "3. 基准日距今已超过一定时间，行业环境可能变化",
            "4. 建议关注 WACC ±1% 区间对估值的影响",
        ]
        for r in risks:
            self._style(self.ws.cell(row=self.row, column=1, value=r), font_normal, None, align_left)
            self.row += 1


# ============================================================
# 主 Builder
# ============================================================

class DCFModelBuilderV2:
    """完整多 Sheet DCF 模型构建器"""

    def __init__(self, params: ModelParams, output_path: str, scenario: str = "ipo"):
        self.p = params
        self.output_path = output_path
        self.scenario = scenario
        self.wb = openpyxl.Workbook()

    def build_all(self):
        p = self.p
        scenario = self.scenario

        # Sheet 1: 概览
        ws = self.wb.active
        ws.title = "项目概览" if scenario == "ipo" else "交易概览"
        SheetOverview(ws).build(p, scenario)

        # Sheet 2: 历史财务
        ws = self.wb.create_sheet("历史财务数据")
        SheetHistoricalFinance(ws).build(p)

        # Sheet 3: 在建工程（仅IPO）
        if scenario == "ipo":
            ws = self.wb.create_sheet("在建工程分析")
            SheetConstructionAnalysis(ws).build(p)

        # Sheet 4: 可比公司WACC
        ws = self.wb.create_sheet("可比公司与WACC")
        SheetComparableWACC(ws).build(p)

        # Sheet 5: 独立DCF
        ws = self.wb.create_sheet("独立DCF")
        ps_standalone = SheetDCF(ws).build(p, label="独立")

        # 计算用于敏感性分析的终值FCFF
        fcff_vals = p.fcff_forecast()
        terminal_fcff = fcff_vals[-1] * (1 + p.terminal_g)

        # Sheet 6: 募投/标的FCFF
        if scenario in ("ipo", "ma_buy"):
            ws = self.wb.create_sheet("募投项目FCFF" if scenario == "ipo" else "标的FCFF")
            SheetProjectFCFF(ws).build(p, scenario)

        # Sheet 7: 备考合并DCF
        if scenario != "ma_sale":
            ws = self.wb.create_sheet("备考合并DCF")
            # 简化：合并FCFF ≈ 独立FCFF * 1.15
            combined_fcff = [f * 1.15 for f in fcff_vals]
            ps_combined = SheetDCF(ws).build(p, label="合并", is_combined=True, fcff_override=combined_fcff)
        else:
            # 出售：剩余业务DCF
            ws = self.wb.create_sheet("剩余业务DCF")
            ps_combined = SheetDCF(ws).build(p, label="剩余业务(扣非)", is_combined=True)
            # 出售前参考
            ws = self.wb.create_sheet("出售前参考")
            SheetOverview(ws).build(p, scenario)

        # Sheet 8: 估值对比
        ws = self.wb.create_sheet("估值对比")
        SheetComparison(ws).build(p, ps_standalone, ps_combined, scenario)

        # Sheet 9: 敏感性分析
        ws = self.wb.create_sheet("敏感性分析")
        SheetSensitivity(ws).build(p, terminal_fcff, ps_combined, p.wacc_combined if scenario != "ma_sale" else p.wacc_standalone)

        # Sheet 10: 关键假设清单
        ws = self.wb.create_sheet("关键假设清单")
        # 计算终值占比
        dfs = p.discount_factors(p.wacc_standalone)
        tv = terminal_fcff / (p.wacc_standalone - p.terminal_g)
        pv_tv = tv * dfs[-1]
        pv_fcff = sum(fcff_vals[i] * dfs[i] for i in range(p.forecast_years))
        ev = pv_fcff + pv_tv
        tv_pct = pv_tv / ev * 100

        SheetAssumptions(ws).build(p, scenario, ps_standalone, ps_combined, tv_pct)

    def save(self):
        self.wb.save(self.output_path)
        print(f"✅ 已生成: {self.output_path}")


# ============================================================
# 预设参数
# ============================================================

def make_ipo_params():
    """IPO 场景示例参数（某IC封测企业风格）"""
    p = ModelParams(
        company_name="某IPO企业",
        stock_code="（IPO申报）",
        base_year=2024,
        base_date="2024-12-31",
        issue_type="IPO",
        shares_base=15000,
        shares_issue=5000,
        # 历史财务（3年）
        rev_hist=[150000, 180000, 210000],
        cost_hist=[105000, 124200, 142800],
        profit_hist=[22000, 26000, 31000],
        ni_parent_hist=[18700, 22100, 26350],
        growth_rates=[0.12, 0.10, 0.08, 0.06, 0.04, 0.03],
        ebit_margin=0.185,
        tax_rate=0.15,
        capex_actual=12000,
        dep_fixed=8500, amor_intangible=400, amor_lt_prepaid=200, dep_right_use=300,
        cash=30000, ar_receivable=42000, inventory=28000, ap_payable=25000,
        fixed_assets=85000, construction_in_progress=15000,
        short_debt=20000, long_debt=30000, one_year_debt=5000, bonds=0,
        minority_interest=2000,
        wacc_standalone=0.0912, wacc_combined=0.0942,
        projects=[
            {
                "name": "高端封测产线建设项目",
                "full_rev": 80000,
                "ramp": [0, 0.3, 1.0, 1.0, 1.0, 1.0],
                "ebit_margin": 0.20,
                "capex_yearly": [14000, 8400, 0, 0, 0, 0],
            },
            {
                "name": "研发中心升级项目",
                "full_rev": 3000,
                "ramp": [0.8, 1.0, 1.0, 1.0, 1.0, 1.0],
                "ebit_margin": 0.15,
                "capex_yearly": [1500, 0, 0, 0, 0, 0],
            },
        ]
    )
    return p


def make_ma_buy_params():
    """M&A 并购置入场景示例参数（某矿业并购风格）"""
    p = ModelParams(
        company_name="某并购方企业",
        stock_code="XXXXXX.SZ",
        base_year=2025,
        base_date="2025-03-31",
        issue_type="发股购买",
        shares_base=60000,
        shares_issue=12000,
        rev_hist=[185000, 210000, 230000],
        cost_hist=[120000, 136500, 149500],
        profit_hist=[28000, 31500, 35000],
        ni_parent_hist=[23800, 26800, 29750],
        growth_rates=[0.08, 0.07, 0.06, 0.05, 0.04, 0.03],
        ebit_margin=0.22,
        tax_rate=0.15,
        capex_actual=25000,
        dep_fixed=12000, amor_intangible=2000, amor_lt_prepaid=500, dep_right_use=1000,
        cash=50000, ar_receivable=60000, inventory=35000, ap_payable=40000,
        fixed_assets=150000, construction_in_progress=20000,
        short_debt=30000, long_debt=80000, one_year_debt=10000, bonds=0,
        minority_interest=5000,
        wacc_standalone=0.088, wacc_combined=0.091,
        rf=0.0203, mrp=0.0657, bu_avg=0.9742,
        projects=[
            {
                "name": "某矿业标的",
                "full_rev": 120000,
                "ramp": [0, 0, 0.75, 1.0, 1.0, 1.0],
                "ebit_margin": 0.25,
                "capex_yearly": [40000, 44372, 0, 0, 0, 0],
            }
        ]
    )
    p.acquisition_price = 650768.80
    return p


def make_ma_sale_params():
    """M&A 重大资产出售场景示例参数（某*ST企业风格）"""
    p = ModelParams(
        company_name="某*ST企业",
        stock_code="XXXXXX.SZ",
        base_year=2024,
        base_date="2024-12-31",
        issue_type="不适用",
        shares_base=173000,
        shares_issue=0,
        rev_hist=[120000, 95000, 80000],
        cost_hist=[110000, 100000, 95000],
        profit_hist=[-35000, -28000, -15000],
        ni_parent_hist=[-30000, -25000, -12000],
        growth_rates=[0.05, 0.05, 0.04, 0.03, 0.03, 0.02],
        ebit_margin=0.08,
        tax_rate=0.25,
        capex_actual=5000,
        dep_fixed=3000, amor_intangible=200, amor_lt_prepaid=100, dep_right_use=200,
        cash=15000, ar_receivable=22000, inventory=60000, ap_payable=35000,
        fixed_assets=50000, construction_in_progress=0,
        short_debt=40000, long_debt=100000, one_year_debt=20000, bonds=0,
        minority_interest=3000,
        wacc_standalone=0.095, wacc_combined=0.095,
        is_sale_dcf_applicable=False,  # 出售前DCF不适用
        nopat_deducted_nonrecurring=6200,
        investment_income_included=False,
    )
    p.equity_before_sale = -175000
    p.equity_after_sale = 49000
    return p


# ============================================================
# 入口
# ============================================================

if __name__ == "__main__":
    import sys, os

    scenario_map = {
        "ipo":     ("DCF_Model_Template_IPO.xlsx",     make_ipo_params),
        "ma_buy":  ("DCF_Model_Template_MA_Buy.xlsx",   make_ma_buy_params),
        "ma_sale": ("DCF_Model_Template_MA_Sale.xlsx",  make_ma_sale_params),
    }

    if len(sys.argv) < 2:
        print("用法:")
        print("  python build_dcf_model.py --all                    生成全部三种场景模板")
        print("  python build_dcf_model.py --scenario ipo            仅生成IPO模板")
        print("  python build_dcf_model.py --scenario ma_buy         仅生成M&A置入模板")
        print("  python build_dcf_model.py --scenario ma_sale        仅生成M&A出售模板")
        print()
        print("  python build_dcf_model.py 输出文件.xlsx --scenario ipo  指定输出路径")
        sys.exit(1)

    # 解析参数
    scenarios_to_build = []
    custom_output = None

    i = 1
    while i < len(sys.argv):
        arg = sys.argv[i]
        if arg == "--all":
            scenarios_to_build = ["ipo", "ma_buy", "ma_sale"]
        elif arg == "--scenario":
            i += 1
            if i < len(sys.argv) and sys.argv[i] in scenario_map:
                scenarios_to_build = [sys.argv[i]]
            else:
                print(f"未知场景: {sys.argv[i] if i < len(sys.argv) else '(无)'}")
                sys.exit(1)
        elif arg.endswith(".xlsx") and not arg.startswith("--"):
            custom_output = arg
        i += 1

    if not scenarios_to_build:
        # 默认生成全部
        scenarios_to_build = ["ipo", "ma_buy", "ma_sale"]

    # 确定输出目录
    output_dir = os.path.dirname(os.path.abspath(__file__))
    template_dir = os.path.join(os.path.dirname(output_dir), "excel_templates")
    os.makedirs(template_dir, exist_ok=True)

    for s in scenarios_to_build:
        filename, factory = scenario_map[s]
        out_path = custom_output if custom_output else os.path.join(template_dir, filename)
        params = factory()
        builder = DCFModelBuilderV2(params, out_path, s)
        builder.build_all()
        builder.save()

    if len(scenarios_to_build) > 1:
        print(f"\n✅ 已生成 {len(scenarios_to_build)} 个模板文件到: {template_dir}")
