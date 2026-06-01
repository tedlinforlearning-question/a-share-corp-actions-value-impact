#!/usr/bin/env python3
"""生成再融资估值分析空模板（纯结构，无案例数据）。"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 样式定义 ──
TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=12)
HEADER_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11)
BODY_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
SUBTOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def setup_sheet(ws, title, headers):
    """通用Sheet设置：标题行 + 表头"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Freeze top rows
    ws.freeze_panes = "A3"

    # Auto width
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(16, len(headers[col_idx - 1]) * 2 + 4)


# ═══════════════════════════════════════════════════════════════
# 模板1: 纯补流 7 Sheet
# ═══════════════════════════════════════════════════════════════
def build_pure_wc_template(path):
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 参数汇总
    s1 = wb.create_sheet("1-参数汇总")
    setup_sheet(s1, "参数汇总 — 发行方案与核心财务数据", ["参数项", "取值", "单位", "数据来源", "备注"])
    rows_s1 = [
        ["发行前总股本", "", "万股", "募集说明书·股权结构", ""],
        ["发行数量上限", "", "万股", "募集说明书·发行方案", "≤总股本30%"],
        ["发行价格(底价)", "", "元/股", "募集说明书", ""],
        ["募集资金总额(上限)", "", "万元", "计算=数量×价格", ""],
        ["募集资金净额", "", "万元", "募集说明书", "扣除发行费用"],
        ["", "", "", "", ""],
        ["归母净资产(发行前)", "", "万元", "审计报告/募集说明书", ""],
        ["每股净资产BPS(发行前)", "", "元/股", "计算=归母净资产/总股本", ""],
        ["资产负债率", "", "%", "募集说明书", ""],
        ["有息负债(发行前)", "", "万元", "审计报告", "短期+长期+一年内到期"],
        ["利息支出(年)", "", "万元", "审计报告附注", ""],
        ["平均借款利率", "", "%", "计算或募集说明书", ""],
        ["有效所得税率", "", "%", "最近一年实际", "所得税/利润总额"],
    ]
    for i, row in enumerate(rows_s1, 3):
        for j, v in enumerate(row, 1):
            c = s1.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 2: 持股比例摊薄
    s2 = wb.create_sheet("2-持股比例摊薄")
    setup_sheet(s2, "持股比例摊薄分析", ["项目", "发行前", "发行后", "变化", "备注"])
    rows_s2 = [
        ["总股本(万股)", "", "", "", ""],
        ["新增股本(万股)", "", "", "", ""],
        ["稀释率", "", "", "", "=新增/发行后"],
        ["", "", "", "", ""],
        ["【具体持股方1】持股(万股)", "", "", "", ""],
        ["  持股比例", "", "", "", ""],
        ["【具体持股方2】持股(万股)", "", "", "", ""],
        ["  持股比例", "", "", "", ""],
    ]
    for i, row in enumerate(rows_s2, 3):
        for j, v in enumerate(row, 1):
            c = s2.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 3: BPS变化
    s3 = wb.create_sheet("3-BPS变化")
    setup_sheet(s3, "每股净资产(BPS)变化分析", ["项目", "发行前", "发行后", "变化", "备注"])
    rows_s3 = [
        ["归母净资产(万元)", "", "", "", ""],
        ["募集资金净额(万元)", "", "", "", ""],
        ["总股本(万股)", "", "", "", ""],
        ["BPS(元/股)", "", "", "", "=归母净资产/总股本"],
        ["", "", "", "", ""],
        ["发行价 vs BPS前", "", "", "", ""],
        ["结论：增厚/摊薄？", "", "", "", "发行价>BPS前→增厚(老股东受益)"],
    ]
    for i, row in enumerate(rows_s3, 3):
        for j, v in enumerate(row, 1):
            c = s3.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 4: EPS摊薄
    s4 = wb.create_sheet("4-EPS摊薄")
    setup_sheet(s4, "EPS摊薄分析 — 三情景", ["项目", "悲观情景", "基准情景", "乐观情景", "备注"])
    rows_s4 = [
        ["归母净利润假设(万元)", "", "", "", ""],
        ["总估值(万元)", "", "", "", "P/E法或DCF"],
        ["发行后总股本(万股)", "", "", "", ""],
        ["EPS(元/股)", "", "", "", "=净利润/总股本"],
        ["每股内在价值(元/股)", "", "", "", "=总估值/总股本"],
        ["", "", "", "", ""],
        ["发行价", "", "", "", ""],
        ["发行价 vs EPS", "", "", "", ""],
        ["发行价 vs 内在价值", "", "", "", ""],
    ]
    for i, row in enumerate(rows_s4, 3):
        for j, v in enumerate(row, 1):
            c = s4.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 5: 利息节约
    s5 = wb.create_sheet("5-利息节约")
    setup_sheet(s5, "利息节约测算", ["项目", "数值", "单位", "备注"])
    rows_s5 = [
        ["募集资金净额", "", "万元", ""],
        ["其中补流部分", "", "万元", ""],
        ["拟偿还借款金额", "", "万元", ""],
        ["平均借款利率", "", "%", ""],
        ["税前利息节约", "", "万元/年", "=偿还金额×利率"],
        ["所得税率", "", "%", ""],
        ["税后利息节约", "", "万元/年", "=税前×(1-T)"],
        ["", "", "", ""],
        ["利润改善幅度", "", "%", "=税后节约/税前利润"],
    ]
    for i, row in enumerate(rows_s5, 3):
        for j, v in enumerate(row, 1):
            c = s5.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 6: Pro Forma 资产负债表
    s6 = wb.create_sheet("6-ProForma资产负债表")
    setup_sheet(s6, "Pro Forma 合并资产负债表(发行前后对比)", ["项目", "发行前(万元)", "发行调整(万元)", "发行后(万元)", "备注"])
    rows_s6 = [
        ["资产", "", "", "", ""],
        ["  货币资金", "", "", "=前+募资净额", ""],
        ["  应收账款(含票据)", "", "", "", ""],
        ["  存货", "", "", "", ""],
        ["  固定资产", "", "", "", ""],
        ["  在建工程", "", "", "", ""],
        ["  总资产", "", "", "", ""],
        ["负债", "", "", "", ""],
        ["  有息负债", "", "", "=前-偿还借款", ""],
        ["  应付账款", "", "", "", ""],
        ["  总负债", "", "", "", ""],
        ["权益", "", "", "", ""],
        ["  归母权益", "", "", "=前+募资净额", ""],
        ["  少数股东权益", "", "", "", ""],
    ]
    for i, row in enumerate(rows_s6, 3):
        for j, v in enumerate(row, 1):
            c = s6.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 7: 综合对比
    s7 = wb.create_sheet("7-综合对比")
    setup_sheet(s7, "综合对比总览", ["维度", "发行前", "发行后", "变化", "评判"])
    rows_s7 = [
        ["总股本(万股)", "", "", "", ""],
        ["稀释率", "", "", "", ""],
        ["BPS(元/股)", "", "", "", ""],
        ["资产负债率", "", "", "", ""],
        ["有息负债(万元)", "", "", "", ""],
        ["利息节约(万元/年)", "", "", "", ""],
        ["EPS基准情景(元/股)", "", "", "", ""],
        ["", "", "", "", ""],
        ["最终评判", "", "", "", "【增厚/摊薄判断 + 综合建议】"],
    ]
    for i, row in enumerate(rows_s7, 3):
        for j, v in enumerate(row, 1):
            c = s7.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    wb.save(path)
    print(f"✅ 纯补流(7 Sheet)模板 → {path}")


# ═══════════════════════════════════════════════════════════════
# 模板2: 补流+募投 10 Sheet
# ═══════════════════════════════════════════════════════════════
def build_wc_project_template(path):
    wb = Workbook()
    wb.remove(wb.active)

    # Sheets 1-7 same as pure WC template (copy structure)
    for sheet_def in [
        ("1-参数汇总", "参数汇总 — 发行方案+核心财务数据+募投效益测算",
         ["参数项", "取值", "单位", "数据来源", "备注"]),
        ("2-持股比例摊薄", "持股比例摊薄分析",
         ["项目", "发行前", "发行后", "变化", "备注"]),
        ("3-BPS变化", "每股净资产(BPS)变化",
         ["项目", "发行前", "发行后", "变化", "备注"]),
        ("4-EPS摊薄", "EPS摊薄分析",
         ["项目", "悲观", "基准", "乐观", "备注"]),
        ("5-利息节约", "利息节约测算",
         ["项目", "数值", "单位", "备注"]),
        ("6-ProForma", "Pro Forma资产负债表",
         ["项目", "发行前(万元)", "调整(万元)", "发行后(万元)", "备注"]),
        ("7-综合对比", "综合对比总览",
         ["维度", "发行前", "发行后", "变化", "评判"]),
    ]:
        ws = wb.create_sheet(sheet_def[0])
        setup_sheet(ws, sheet_def[1], sheet_def[2])

    # Sheet 8: 项目一DCF
    s8 = wb.create_sheet("8-项目一DCF")
    setup_sheet(s8, "项目一 DCF 估值模型", ["项目", "建设期", "达产期Y1", "Y2", "Y3", "Y4", "Y5", "Y6", "Y7", "公式/来源"])
    rows_proj = [
        ["营业收入(万元)", "", "", "", "", "", "", "", "", "募集说明书效益测算表"],
        ["营业成本(万元)", "", "", "", "", "", "", "", "", ""],
        ["利润总额(万元)", "", "", "", "", "", "", "", "", ""],
        ["所得税(万元)", "", "", "", "", "", "", "", "", ""],
        ["净利润(万元)", "", "", "", "", "", "", "", "", ""],
        ["折旧摊销(万元)", "", "", "", "", "", "", "", "", ""],
        ["CAPEX(万元)", "", "", "", "", "", "", "", "", ""],
        ["FCF(万元)", "", "", "", "", "", "", "", "", "=NOPAT+D&A-CAPEX"],
        ["", "", "", "", "", "", "", "", "", ""],
        ["WACC", "", "", "", "", "", "", "", "", ""],
        ["永续增长率g", "", "", "", "", "", "", "", "", ""],
        ["NPV(万元)", "", "", "", "", "", "", "", "", "=PV(预测FCF)+PV(终值)"],
        ["IRR", "", "", "", "", "", "", "", "", "IRR(FCFF)=?"],
    ]
    for i, row in enumerate(rows_proj, 3):
        for j, v in enumerate(row, 1):
            c = s8.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 9: 项目二DCF
    s9 = wb.create_sheet("9-项目二DCF")
    setup_sheet(s9, "项目二 DCF 估值模型(同项目一)", ["项目", "建设期", "达产期Y1", "Y2", "Y3", "Y4", "Y5", "Y6", "Y7", "公式/来源"])
    for i, row in enumerate(rows_proj, 3):
        for j, v in enumerate(row, 1):
            c = s9.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 10: 合并估值
    s10 = wb.create_sheet("10-合并估值")
    setup_sheet(s10, "合并估值 — 存量P/E + Σ项目NPV", ["项目", "数值", "单位", "每股贡献(元)", "备注"])
    rows_summary = [
        ["发行前总股本", "", "万股", "", ""],
        ["新增股本", "", "万股", "", ""],
        ["发行后总股本", "", "万股", "", ""],
        ["", "", "", "", ""],
        ["存量业务估值(P/E或DCF)", "", "万元", "", ""],
        ["项目一NPV", "", "万元", "", ""],
        ["项目二NPV", "", "万元", "", ""],
        ["发行后总价值", "", "万元", "", "=存量+ΣNPV"],
        ["", "", "", "", ""],
        ["发行前每股价值", "", "元/股", "", "=存量估值/发行前股本"],
        ["发行后每股价值", "", "元/股", "", "=总价值/发行后股本"],
        ["股本稀释效应", "", "元/股", "", "=存量发行后每股-发行前每股"],
        ["项目创造价值", "", "元/股", "", "=ΣNPV/发行后股本"],
        ["净效应", "", "元/股", "", "=发行后每股-发行前每股"],
        ["", "", "", "", ""],
        ["判断：项目NPV能否覆盖稀释？", "", "", "", "净效应>0→覆盖✓"],
    ]
    for i, row in enumerate(rows_summary, 3):
        for j, v in enumerate(row, 1):
            c = s10.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    wb.save(path)
    print(f"✅ 补流+募投(10 Sheet)模板 → {path}")


# ═══════════════════════════════════════════════════════════════
# 模板3: 收购+配套 7 Sheet
# ═══════════════════════════════════════════════════════════════
def build_acquisition_template(path):
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 交易概览
    s1 = wb.create_sheet("1-交易概览")
    setup_sheet(s1, "交易概览 — 收购资产+配套融资结构", ["项目", "数值", "单位", "数据来源", "备注"])
    rows_txn = [
        ["收购标的", "", "", "募集说明书", "标的公司名称/股权比例"],
        ["收购对价", "", "万元", "募集说明书", ""],
        ["标的评估方法", "", "", "评估报告", "资产基础法/收益法"],
        ["标的评估值", "", "万元", "评估报告", ""],
        ["", "", "", "", ""],
        ["发行前总股本", "", "万股", "募集说明书", ""],
        ["发行数量上限", "", "万股", "募集说明书", ""],
        ["发行价格(底价)", "", "元/股", "募集说明书", ""],
        ["募集资金总额", "", "万元", "计算", ""],
        ["", "", "", "", ""],
        ["发行人最近年度收入", "", "万元", "年报/审计报告", ""],
        ["发行人归母净利润", "", "万元", "年报/审计报告", ""],
        ["发行人ROE", "", "%", "计算", ""],
    ]
    for i, row in enumerate(rows_txn, 3):
        for j, v in enumerate(row, 1):
            c = s1.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 2: 摊薄分析
    s2 = wb.create_sheet("2-摊薄分析")
    setup_sheet(s2, "持股比例 & BPS变化", ["项目", "发行前", "发行后", "变化", "备注"])
    rows_dilution = [
        ["总股本(万股)", "", "", "", ""],
        ["新增股本(万股)", "", "", "", ""],
        ["稀释率", "", "", "", ""],
        ["", "", "", "", ""],
        ["归母净资产(万元)", "", "", "", ""],
        ["BPS(元/股)", "", "", "", ""],
        ["发行价 vs BPS前", "", "", "", "增厚/摊薄判断"],
    ]
    for i, row in enumerate(rows_dilution, 3):
        for j, v in enumerate(row, 1):
            c = s2.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 3: 标的财务
    s3 = wb.create_sheet("3-标的财务分析")
    setup_sheet(s3, "标的公司财务分析", ["项目", "最近1年", "最近2年", "单位", "备注"])
    rows_target = [
        ["营业收入", "", "", "万元", ""],
        ["营业成本", "", "", "万元", ""],
        ["营业利润", "", "", "万元", ""],
        ["净利润", "", "", "万元", ""],
        ["", "", "", "", ""],
        ["总资产", "", "", "万元", ""],
        ["净资产", "", "", "万元", ""],
        ["", "", "", "", ""],
        ["P/E(收购对价/净利润)", "", "", "倍", ""],
        ["P/B(收购对价/净资产)", "", "", "倍", ""],
        ["", "", "", "", ""],
        ["产能规模", "", "", "", ""],
        ["产销率", "", "", "%", ""],
    ]
    for i, row in enumerate(rows_target, 3):
        for j, v in enumerate(row, 1):
            c = s3.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 4: 标的DCF
    s4 = wb.create_sheet("4-标的DCF")
    setup_sheet(s4, "标的公司独立DCF估值", ["项目", "Y1E", "Y2E", "Y3E", "Y4E", "Y5E", "终年", "公式/来源"])
    rows_dcf = [
        ["营业收入(万元)", "", "", "", "", "", "", "产能×价格假设"],
        ["营业成本(万元)", "", "", "", "", "", "", ""],
        ["EBIT(万元)", "", "", "", "", "", "", ""],
        ["NOPAT(万元)", "", "", "", "", "", "", "=EBIT×(1-T)"],
        ["D&A(万元)", "", "", "", "", "", "", ""],
        ["CAPEX(万元)", "", "", "", "", "", "", ""],
        ["ΔNWC(万元)", "", "", "", "", "", "", ""],
        ["FCFF(万元)", "", "", "", "", "", "", "=NOPAT+D&A-CAPEX-ΔNWC"],
        ["", "", "", "", "", "", "", ""],
        ["折现因子", "", "", "", "", "", "", ""],
        ["PV(FCFF)(万元)", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", ""],
        ["WACC", "", "", "", "", "", "", "标的行业β×市场参数"],
        ["永续增长率g", "", "", "", "", "", "", ""],
        ["终值TV(万元)", "", "", "", "", "", "", ""],
        ["EV(万元)", "", "", "", "", "", "", "=ΣPV(FCFF)+PV(TV)"],
        ["+现金(万元)", "", "", "", "", "", "", ""],
        ["-有息负债(万元)", "", "", "", "", "", "", ""],
        ["-少数股东权益(万元)", "", "", "", "", "", "", ""],
        ["股权价值(万元)", "", "", "", "", "", "", ""],
        ["收购比例对应价值(万元)", "", "", "", "", "", "", "=股权价值×收购%"],
    ]
    for i, row in enumerate(rows_dcf, 3):
        for j, v in enumerate(row, 1):
            c = s4.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 5: 备考合并
    s5 = wb.create_sheet("5-备考合并")
    setup_sheet(s5, "备考合并视角(含关联交易抵消)", ["项目", "发行人", "标的", "抵消/调整", "备考合并", "备注"])
    rows_consol = [
        ["营业收入(万元)", "", "", "", "", "标的内部销售需抵消"],
        ["营业成本(万元)", "", "", "", "", "发行人采购标的金额抵消"],
        ["毛利(万元)", "", "", "", "", "=收入-成本"],
        ["毛利率", "", "", "", "", "合并后通常上升"],
        ["", "", "", "", "", ""],
        ["归母净利润(万元)", "", "", "", "", "扣除少数股东"],
        ["总资产(万元)", "", "", "", "", ""],
        ["归母净资产(万元)", "", "", "", "", ""],
    ]
    for i, row in enumerate(rows_consol, 3):
        for j, v in enumerate(row, 1):
            c = s5.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 6: 协同效应
    s6 = wb.create_sheet("6-协同效应")
    setup_sheet(s6, "协同效应量化", ["协同类型", "年化节约(万元)", "兑现时间", "永续价值(万元)", "数据来源/假设"])
    rows_synergy = [
        ["① 成本节约(原材料自产替代外购)", "", "", "", "价差×年用量"],
        ["   - 原材料A价差节约", "", "", "", ""],
        ["   - 运输/物流成本降低", "", "", "", ""],
        ["② 收入增长(交叉销售)", "", "", "", "如可量化"],
        ["③ 管理整合(职能合并)", "", "", "", "一般定性描述"],
        ["", "", "", "", ""],
        ["年化协同合计(万元)", "", "", "", ""],
        ["协同永续价值(万元)", "", "", "", "=年化合计/WACC"],
        ["", "", "", "", ""],
        ["协同/收购对价", "", "", "", ">100%→极度划算"],
        ["协同/收购对价", "", "", "", "50-100%→合理"],
    ]
    for i, row in enumerate(rows_synergy, 3):
        for j, v in enumerate(row, 1):
            c = s6.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    # Sheet 7: 价值总结
    s7 = wb.create_sheet("7-价值影响总结")
    setup_sheet(s7, "综合价值影响总结", ["项目", "金额(万元)", "每股(元/股)", "占比", "备注"])
    rows_val = [
        ["发行人独立价值", "", "", "", "P/E或DCF"],
        ["标的DCF股权价值(收购比例)", "", "", "", ""],
        ["协同永续价值", "", "", "", ""],
        ["", "", "", "", ""],
        ["发行前总价值", "", "", "", "=发行人价值"],
        ["发行后总价值", "", "", "", "=发行人+标的+协同"],
        ["", "", "", "", ""],
        ["① 股本稀释效应(每股)", "", "", "", "负值(发行后-前)"],
        ["② 标的贡献(每股)", "", "", "", "标的价值/发行后股本"],
        ["③ 协同贡献(每股)", "", "", "", "协同价值/发行后股本"],
        ["净效应(每股)", "", "", "", "=②+③+①"],
        ["", "", "", "", ""],
        ["判断：收购是否创造每股价值？", "", "", "", "净效应>0→创造✓"],
    ]
    for i, row in enumerate(rows_val, 3):
        for j, v in enumerate(row, 1):
            c = s7.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    wb.save(path)
    print(f"✅ 收购+配套(7 Sheet)模板 → {path}")


# ── 主入口 ──
if __name__ == "__main__":
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    templates_dir = os.path.join(base, "excel_templates")
    os.makedirs(templates_dir, exist_ok=True)

    build_pure_wc_template(os.path.join(templates_dir, "Refinancing_WC_Dilution_Template.xlsx"))
    build_wc_project_template(os.path.join(templates_dir, "Refinancing_WC_Project_DCF_Template.xlsx"))
    build_acquisition_template(os.path.join(templates_dir, "Refinancing_Acquisition_Template.xlsx"))
    print("\n🎉 全部再融资模板生成完毕。")
